import pandas as pd
import re
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.page import PageMargins

def get_sort_keys(address):
    addr = str(address).strip()
    if addr.lower() == 'nan' or not addr:
        return pd.Series(['', 0, '', 0])

    house_num = 0
    house_letter = ''
    street_part = addr

    house_match = re.search(r'(буд\.|б\.|буд\s|будинок)\s*(\d+)\s*([А-Яа-яІіЇїЄєA-Za-z]?)', addr, re.IGNORECASE)
    if house_match:
        house_num = int(house_match.group(2))
        house_letter = house_match.group(3).upper()
        street_part = addr[:house_match.start()]

    flat_num = 0
    flat_match = re.search(r'(кв\.|квартира|кв\s)\s*(\d+)', addr, re.IGNORECASE)
    if flat_match:
        flat_num = int(flat_match.group(2))

    sort_street = re.sub(r'[\s\,\.\-\_]', '', street_part.lower())
    return pd.Series([sort_street, house_num, house_letter, flat_num])

def run_routes_generation(input_path, output_dir):
    """
    Генерує окремі відомості для кожного відповідального.
    Повертає список словників з інформацією про файли.
    """
    session_id = str(uuid.uuid4())[:6]
    generated_files = []

    # Читаємо файл (header=1, як у тебе)
    df = pd.read_excel(input_path, header=1, dtype=str)
    df = df.dropna(how='all').fillna('').replace('nan', '')

    def find_col(keyword):
        for col in df.columns:
            if keyword.lower() in str(col).lower(): return col
        return None

    col_or = find_col("номер ор") or find_col("ор")
    col_name = find_col("контрагент")
    col_phone = find_col("телефон")
    col_date = find_col("дата")
    col_resp = find_col("відповідальний")
    col_address = find_col("адрес")
    col_meter = find_col("лічильника")
    col_pokaz = find_col("покази")
    col_debt = find_col("поточний борг") or find_col("борг") or find_col("сума")

    if not col_address: raise ValueError("Колонка з адресою не знайдена!")

    # Сортування (твоя магія)
    df[['Sort_Street', 'Sort_House', 'Sort_Letter', 'Sort_Flat']] = df[col_address].apply(get_sort_keys)
    df = df.sort_values(by=['Sort_Street', 'Sort_House', 'Sort_Letter', 'Sort_Flat'])

    grouped = df.groupby(col_resp)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    for person, group_df in grouped:
        if person == '' or str(person).lower() == 'nan': continue

        # Генеруємо унікальне ім'я файлу для завантаження
        safe_person_name = str(person).replace('/', '_').replace('\\', '_')
        filename = f"{session_id}_{safe_person_name}.xlsx"
        file_path = os.path.join(output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Відомість"

        headers = ["№ з/п", "Номер ОР", "Контрагент", "Телефон", "Дата", "Відповідальний за вручення", "Адреси точок обліку", "Номер лічильника", "Останні покази", "Сума поточного боргу", "Факт показ"]
        ws.append(headers)

        # Дизайн шапки
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # Заповнення даними
        row_idx = 2
        numbering = 1
        for _, row_data in group_df.iterrows():
            addr_val = str(row_data[col_address]).strip()
            if not addr_val or addr_val.lower() == 'nan': continue

            val_debt = str(row_data[col_debt]).strip() if col_debt else ""
            if val_debt.endswith('.0'): val_debt = val_debt[:-2]
            val_or = str(row_data[col_or]).strip() if col_or else ""
            if val_or.endswith('.0'): val_or = val_or[:-2]

            new_row = [numbering, val_or, str(row_data[col_name]).strip(), str(row_data[col_phone]).strip(), str(row_data[col_date]).strip(), str(row_data[col_resp]).strip(), addr_val, str(row_data[col_meter]).strip(), str(row_data[col_pokaz]).strip(), val_debt, ""]
            ws.append(new_row)

            for c_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.number_format = '@'
                align_h = 'left' if headers[c_idx-1] in ["Адреси точок обліку", "Контрагент"] else 'center'
                cell.alignment = Alignment(horizontal=align_h, vertical='bottom', wrap_text=True)
                cell.border = thin_border
            row_idx += 1
            numbering += 1

        # Твоє форматування сторінки
        widths = {'A': 4, 'B': 14, 'C': 22, 'D': 13, 'E': 10, 'F': 14, 'G': 28, 'H': 12, 'I': 9, 'J': 11, 'K': 12}
        for col_letter, width in widths.items(): ws.column_dimensions[col_letter].width = width

        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.paperSize = 9
        ws.page_setup.orientation = "landscape"
        ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.4, bottom=0.4, header=0.0, footer=0.0)

        wb.save(file_path)
        
        # Додаємо у список результатів
        generated_files.append({"label": str(person), "filename": filename})

    return generated_files