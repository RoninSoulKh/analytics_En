import os
import re
import uuid
import pandas as pd
import numpy as np
import calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def run_calls_analysis(input_path: str, output_dir: str) -> str:
    # --- СТИЛІ З КОЛАБУ ---
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    bold_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, horizontal='center', vertical='center')
    
    # Додамо трохи стилів для Зведеної статистики (шапка та футер)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="50A74B", end_color="50A74B", fill_type="solid")
    controller_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

    # 1. Читаємо дані через pandas для математики та сортування
    df = pd.read_excel(input_path, sheet_name=0, header=None)
    data_df = df.iloc[3:].copy()
    
    # ФІКС: Видаляємо "фантомні" порожні рядки, щоб не було збоїв
    data_df = data_df.dropna(subset=[1])
    data_df = data_df.where(pd.notnull(data_df), None)

    # Сортування: ПІБ (колонка 1), Дата (колонка 5)
    data_df[5] = pd.to_datetime(data_df[5], errors='coerce', dayfirst=True)
    data_df = data_df.sort_values(by=[1, 5])

    # 2. Завантажуємо ОРИГІНАЛЬНИЙ файл
    # ВИКОРИСТОВУЄМО ТВОЮ ЛОГІКУ: Не видаляємо лист, а чистимо рядки!
    # Це збереже складну шапку і всі формати перших трьох рядків.
    wb = load_workbook(input_path)
    ws1 = wb.worksheets[0]
    ws1.title = "Поіменний (Оброблений)"

    # Видаляємо всі старі дані з 4-го рядка і до кінця
    max_r = ws1.max_row
    if max_r >= 4:
        ws1.delete_rows(4, max_r - 3)

    ws1.freeze_panes = 'A4'

    rows_data = data_df.values.tolist()

    if rows_data:
        current_name = rows_data[0][1]
        current_date = rows_data[0][5]
    else:
        current_name = None
        current_date = None

    count = 0

    # Вписуємо нові відсортовані дані (ОРИГІНАЛЬНА ЛОГІКА З КОЛАБУ)
    for row in rows_data:
        name = row[1]
        date = row[5]

        # Змінився працівник
        if name != current_name:
            summary_row = [None] * len(row)
            date_str = current_date.strftime('%d.%m.%Y') if pd.notnull(current_date) else 'Невідома дата'
            summary_row[5] = f"Всього дзвінків за {date_str}: {count}"
            ws1.append(summary_row)

            for cell in ws1[ws1.max_row]:
                cell.fill = yellow_fill
                cell.font = bold_font
                cell.border = thin_border

            ws1.append([None])
            ws1.append([None])
            ws1.append([None])

            current_name = name
            current_date = date
            count = 0

        # Змінилася дата
        elif date != current_date:
            summary_row = [None] * len(row)
            date_str = current_date.strftime('%d.%m.%Y') if pd.notnull(current_date) else 'Невідома дата'
            summary_row[5] = f"Всього дзвінків за {date_str}: {count}"
            ws1.append(summary_row)

            for cell in ws1[ws1.max_row]:
                cell.fill = yellow_fill
                cell.font = bold_font
                cell.border = thin_border

            ws1.append([None])
            ws1.append([None])

            current_date = date
            count = 0

        # Форматуємо рядок даних
        formatted_row = list(row)
        if pd.notnull(formatted_row[5]):
            formatted_row[5] = formatted_row[5].strftime('%d.%m.%Y')
        ws1.append(formatted_row)

        # Малюємо сітку для даних
        for cell in ws1[ws1.max_row]:
            if cell.value is not None:
                cell.border = thin_border

        count += 1

    # Закриваємо останнього
    if rows_data:
        summary_row = [None] * len(rows_data[0])
        date_str = current_date.strftime('%d.%m.%Y') if pd.notnull(current_date) else 'Невідома дата'
        summary_row[5] = f"Всього дзвінків за {date_str}: {count}"
        ws1.append(summary_row)
        for cell in ws1[ws1.max_row]:
            cell.fill = yellow_fill
            cell.font = bold_font
            cell.border = thin_border

    # --- ЛИСТ 2: Зведена статистика (ОСЬ ТУТ СУПЕР-ОПТИМІЗАЦІЯ) ---
    if "Зведена статистика" in wb.sheetnames:
        del wb["Зведена статистика"]
    ws2 = wb.create_sheet(title="Зведена статистика")
    ws2.freeze_panes = 'A2'

    valid_dates = data_df[5].dropna()
    if not valid_dates.empty:
        target_year = int(valid_dates.dt.year.mode()[0])
        target_month = int(valid_dates.dt.month.mode()[0])
        _, num_days = calendar.monthrange(target_year, target_month)
        month_dates = [pd.Timestamp(year=target_year, month=target_month, day=d) for d in range(1, num_days + 1)]
        
        # Замість 900+ циклів pandas рахує все за 1 мілісекунду
        pivot_counts = pd.crosstab(data_df[5], data_df[1])
    else:
        month_dates = []
        pivot_counts = pd.DataFrame()

    workers = sorted([w for w in data_df[1].unique() if w is not None])
    default_rrsc = data_df[0].dropna().iloc[0] if not data_df[0].dropna().empty else "Дані відсутні"

    headers = ["РРСЦ", "Дата"] + workers
    ws2.append(headers)

    # Дизайн шапки
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_align

    # Вставляємо готові дані
    for md in month_dates:
        row_out = [default_rrsc, md.strftime('%d.%m.%Y')]
        for w in workers:
            val = pivot_counts.loc[md, w] if (md in pivot_counts.index and w in pivot_counts.columns) else 0
            row_out.append(int(val))
        ws2.append(row_out)

        for cell in ws2[ws2.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Підсумок внизу
    total_row = ["", "Всього за місяць:"]
    for w in workers:
        total_count = pivot_counts[w].sum() if w in pivot_counts.columns else 0
        total_row.append(int(total_count))
    ws2.append(total_row)

    for c_idx, cell in enumerate(ws2[ws2.max_row], 1):
        cell.font = bold_font
        cell.border = thin_border
        if c_idx > 1:
            cell.fill = controller_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.fill = header_fill

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    for col in range(3, len(headers) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 15

    # 3. Збереження (Адаптовано під FastAPI)
    input_filename = os.path.basename(input_path)
    clean_name = re.sub(r'\s*\(\d+\)', '', input_filename)
    uid = str(uuid.uuid4())[:8] # Додаємо унікальний ID, щоб файли не перезаписували один одного
    output_filename = clean_name.replace('.xlsx', f'_{uid}_Оброблений.xlsx')
    output_path = os.path.join(output_dir, output_filename)

    wb.save(output_path)
    return output_filename