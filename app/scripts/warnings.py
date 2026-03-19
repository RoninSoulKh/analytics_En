import os
import uuid
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def fill_missing_dates(df):
    """Заповнює порожні дати на основі сусідніх квартир (Вулиця + Будинок)."""
    df = df.sort_values(by=['Вулиця', 'Будинок', 'Квартира'])
    df['Дата'] = df.groupby(['Вулиця', 'Будинок'])['Дата'].ffill().bfill()
    return df

def run_warnings_analysis(input_path: str, output_dir: str) -> str:
    try:
        # Пропускаємо перші 3 рядки складної шапки
        df = pd.read_excel(input_path, sheet_name='Поіменний', skiprows=3, header=None)
        
        # Перейменовуємо колонки за індексами
        df.rename(columns={
            0: 'РРСЦ', 1: 'ПІБ', 3: 'Вулиця', 4: 'Будинок', 5: 'Квартира',
            10: 'Дата', 16: 'Вруч_шт', 17: 'Вруч_сума', 22: 'Шпарина_шт', 24: 'Шпарина_сума'
        }, inplace=True)
        
        df = df[['РРСЦ', 'ПІБ', 'Вулиця', 'Будинок', 'Квартира', 'Дата', 'Вруч_шт', 'Вруч_сума', 'Шпарина_шт', 'Шпарина_сума']].copy()
        df.dropna(subset=['ПІБ', 'Вулиця'], inplace=True)
        
        for col in ['Вруч_шт', 'Вруч_сума', 'Шпарина_шт', 'Шпарина_сума']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        df['Дата'] = pd.to_datetime(df['Дата'], errors='coerce', dayfirst=True).dt.strftime('%d.%m.%Y').fillna('-')
        df = fill_missing_dates(df)
        
        rrsc_name = df['РРСЦ'].iloc[0] if not df['РРСЦ'].empty else "Невідомий РРСЦ"

        # --- АГРЕГАЦІЯ ДЛЯ ЛИСТА 3 (Поіменна статистика) ---
        agg_funcs = {
            'Будинок': lambda x: ', '.join(sorted(set(x.dropna().astype(str)))),
            'Дата': lambda x: '\n'.join(sorted(set(x.dropna().astype(str)))),
            'РРСЦ': 'count',
            'Вруч_шт': 'sum', 
            'Вруч_сума': 'sum',
            'Шпарина_шт': 'sum',
            'Шпарина_сума': 'sum'
        }
        
        grouped_df = df.groupby(['ПІБ', 'Вулиця']).agg(agg_funcs).reset_index()
        grouped_df.rename(columns={'РРСЦ': 'Всього_в_роботі'}, inplace=True)
        grouped_df['Адреси'] = grouped_df['Вулиця'] + " " + grouped_df['Будинок']

        # --- АГРЕГАЦІЯ ДЛЯ ЛИСТА 4 (Зведена статистика) ---
        summary_df = grouped_df.groupby('ПІБ').agg({
            'Всього_в_роботі': 'sum',
            'Вруч_шт': 'sum',
            'Шпарина_шт': 'sum',
            'Вруч_сума': 'sum',
            'Шпарина_сума': 'sum'
        }).reset_index()
        summary_df['Всього сплат'] = summary_df['Вруч_сума'] + summary_df['Шпарина_сума']

        # Підрахунок загальних сум для футера Зведеної статистики
        total_in_work = summary_df['Всього_в_роботі'].sum()
        total_in_hands = summary_df['Вруч_шт'].sum()
        total_in_door = summary_df['Шпарина_шт'].sum()
        total_payments = summary_df['Всього сплат'].sum()

        # --- ГЕНЕРАЦІЯ ФАЙЛУ ---
        uid = str(uuid.uuid4())[:8]
        out_filename = f"warnings_stats_{uid}.xlsx"
        out_path = os.path.join(output_dir, out_filename)
        
        import shutil
        shutil.copyfile(input_path, out_path)
        wb = load_workbook(out_path)
        
        # СТИЛІ
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="50A74B", end_color="50A74B", fill_type="solid")
        controller_font = Font(bold=True, size=12)
        controller_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # --- ЛИСТ 3: СТАТИСТИКА РОЗНЕСЕННЯ ---
        ws3 = wb.create_sheet("Стат. рознесення")
        headers = ["РРСЦ", "Дата", "Адреси", "Всього в роботі, шт.", "Вруч. в руки", "Сплати вруч.", "В шпарину", "Сплати в шпарину"]
        ws3.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_num)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center_align, thin_border
        ws3.freeze_panes = 'A2'

        current_row = 2
        for pib, group in grouped_df.groupby('ПІБ'):
            # Шапка контролера
            ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            cell_pib = ws3.cell(row=current_row, column=1, value=pib)
            cell_pib.font, cell_pib.fill, cell_pib.alignment = controller_font, controller_fill, center_align
            for c in range(1, 9): ws3.cell(row=current_row, column=c).border = thin_border
            current_row += 1
            
            # Дані
            start_data_row = current_row
            for _, row_data in group.iterrows():
                ws3.cell(row=current_row, column=1, value=rrsc_name).alignment = center_align
                ws3.cell(row=current_row, column=2, value=row_data['Дата']).alignment = center_align
                ws3.cell(row=current_row, column=3, value=row_data['Адреси']).alignment = left_align
                ws3.cell(row=current_row, column=4, value=row_data['Всього_в_роботі']).alignment = center_align
                ws3.cell(row=current_row, column=5, value=row_data['Вруч_шт']).alignment = center_align
                ws3.cell(row=current_row, column=6, value=row_data['Вруч_сума']).alignment = center_align
                ws3.cell(row=current_row, column=7, value=row_data['Шпарина_шт']).alignment = center_align
                ws3.cell(row=current_row, column=8, value=row_data['Шпарина_сума']).alignment = center_align
                for c in range(1, 9): ws3.cell(row=current_row, column=c).border = thin_border
                current_row += 1
                
            # Підсумки контролера
            ws3.cell(row=current_row, column=1, value="ВСЬОГО:").font = Font(bold=True)
            ws3.cell(row=current_row, column=1).alignment = Alignment(horizontal="right")
            ws3.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            
            for i, col_letter in enumerate(['D', 'E', 'F', 'G', 'H'], start=4):
                ws3.cell(row=current_row, column=i, value=f"=SUM({col_letter}{start_data_row}:{col_letter}{current_row-1})").font = Font(bold=True)
            
            for c in range(1, 9):
                ws3.cell(row=current_row, column=c).border = thin_border
                ws3.cell(row=current_row, column=c).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            
            current_row += 4

        # Ширина колонок Лист 3
        cols_width3 = {'A': 20, 'B': 12, 'C': 60, 'D': 15, 'E': 15, 'F': 15, 'G': 15, 'H': 15}
        for col, width in cols_width3.items(): ws3.column_dimensions[col].width = width

        # --- ЛИСТ 4: ЗВЕДЕНА СТАТИСТИКА ---
        ws4 = wb.create_sheet("Зведена стат.")
        headers_summary = ["РРСЦ", "ПІБ відповідальний", "Всього в роботі", "Всього вручено", "Всього в шпарину", "Всього сплат"]
        ws4.append(headers_summary)
        
        for col_num, header in enumerate(headers_summary, 1):
            cell = ws4.cell(row=1, column=col_num)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center_align, thin_border
        ws4.freeze_panes = 'A2'

        sum_row = 2
        for _, row_data in summary_df.iterrows():
            ws4.append([rrsc_name, row_data['ПІБ'], row_data['Всього_в_роботі'], row_data['Вруч_шт'], row_data['Шпарина_шт'], row_data['Всього сплат']])
            for c in range(1, 7):
                ws4.cell(row=sum_row, column=c).border = thin_border
                ws4.cell(row=sum_row, column=c).alignment = center_align if c != 2 else left_align
            sum_row += 1

        # ФУТЕР ЛИСТА 4 (Загальний підсумок)
        ws4.cell(row=sum_row, column=1, value="ВСЬОГО ПО РРСЦ:").font = Font(bold=True, color="FFFFFF")
        ws4.cell(row=sum_row, column=1).alignment = Alignment(horizontal="right")
        ws4.cell(row=sum_row, column=1).fill = header_fill
        ws4.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=2)
        
        ws4.cell(row=sum_row, column=3, value=total_in_work).font = Font(bold=True)
        ws4.cell(row=sum_row, column=4, value=total_in_hands).font = Font(bold=True)
        ws4.cell(row=sum_row, column=5, value=total_in_door).font = Font(bold=True)
        ws4.cell(row=sum_row, column=6, value=total_payments).font = Font(bold=True)
        
        for c in range(1, 7):
            ws4.cell(row=sum_row, column=c).border = thin_border
            if c > 2:
                ws4.cell(row=sum_row, column=c).alignment = center_align
                ws4.cell(row=sum_row, column=c).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Ширина колонок Лист 4
        cols_width4 = {'A': 20, 'B': 40, 'C': 15, 'D': 15, 'E': 15, 'F': 15}
        for col, width in cols_width4.items(): ws4.column_dimensions[col].width = width

        wb.save(out_path)
        return out_filename
        
    except Exception as e:
        print(f"Warnings Analysis Error: {e}")
        raise e