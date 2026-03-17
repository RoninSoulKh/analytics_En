import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import math # Додали математику для розрахунку висоти

def run_debts_analysis(input_path, output_dir):
    # 1. Читання файлу (ідентично колабу)
    if input_path.endswith('.csv'):
        df = pd.read_csv(input_path, sep=None, engine='python', dtype={'Особовий рахунок': str})
    else:
        df = pd.read_excel(input_path, dtype={'Особовий рахунок': str})

    df.columns = df.columns.str.strip()

    # Змінні колонок
    col_city = 'Населений пункт'
    col_start_debt = 'Борг на початок місяця, грн.'
    col_current_debt = 'Поточний борг, грн.'
    col_p_debt = 'Дебіторська заборгованість до 1 місяця, грн.'
    col_payment_OLD = 'Заборгованість невизнана судом, грн.'
    col_payment_NEW = 'Сплати за останній місяць, грн.'

    if col_city in df.columns:
        df[col_city] = df[col_city].astype(str).str.strip().str.upper()
        df[col_city] = df[col_city].replace(r'\s+', ' ', regex=True)

    # Пошук колонок з періодами
    try:
        start_idx = df.columns.get_loc(col_p_debt)
        end_idx = df.columns.get_loc(col_current_debt) + 1
        numeric_cols_range = df.columns[start_idx:end_idx].tolist()
    except:
        numeric_cols_range = []

    # Чистка чисел
    cols_to_clean = numeric_cols_range + [col_start_debt, col_current_debt]
    if col_payment_OLD in df.columns: cols_to_clean.append(col_payment_OLD)
    for col in cols_to_clean:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.').str.replace(' ', ''), errors='coerce').fillna(0)

    # Розрахунок і перейменування
    if col_payment_OLD in df.columns and col_start_debt in df.columns and col_current_debt in df.columns:
        df[col_payment_OLD] = df[col_start_debt] - df[col_current_debt]
        df = df.rename(columns={col_payment_OLD: col_payment_NEW})
        numeric_cols_range = [col_payment_NEW if x==col_payment_OLD else x for x in numeric_cols_range]

    # Сортування
    if col_city in df.columns and col_current_debt in df.columns:
        df = df.sort_values(by=[col_city, col_current_debt], ascending=[True, False])

    # Групування
    processed_rows = []
    columns_list = df.columns.tolist()
    cols_to_sum = list(numeric_cols_range)
    if col_payment_NEW not in cols_to_sum and col_payment_NEW in df.columns:
        cols_to_sum.append(col_payment_NEW)

    if col_city in df.columns:
        for city, group in df.groupby(col_city, sort=False):
            processed_rows.extend(group.to_dict('records'))
            sum_all = group[cols_to_sum].sum()
            debtors = group[group[col_current_debt] >= 0.01] if col_current_debt in group.columns else group
            sum_debtors = debtors[cols_to_sum].sum()

            r1 = {c: None for c in columns_list}
            r1[col_city] = f"ВСЬОГО {city} (Усі)"
            for c in cols_to_sum: r1[c] = sum_all[c]
            if col_p_debt in sum_all and sum_all[col_p_debt] != 0 and col_payment_NEW in sum_all:
                r1['Телефон'] = f"% Опл: {(sum_all[col_payment_NEW]/sum_all[col_p_debt]*100):.1f}%"

            r2 = {c: None for c in columns_list}
            r2[col_city] = f"ВСЬОГО {city} (Боржники)"
            for c in cols_to_sum: r2[c] = sum_debtors[c]
            if col_p_debt in sum_debtors and sum_debtors[col_p_debt] != 0 and col_payment_NEW in sum_debtors:
                r2['Телефон'] = f"% Опл: {(sum_debtors[col_payment_NEW]/sum_debtors[col_p_debt]*100):.1f}%"

            processed_rows.extend([r1, r2, {c: None for c in columns_list}])
    else:
        processed_rows = df.to_dict('records')

    # ------------ Ім'я файлу ---------------
    try:
        raw_name = str(df.iloc[0, 0]).strip()
        safe_name = "".join([c for c in raw_name if c.isalnum() or c in (' ', '-', '_', '.')]).strip()
        output_name = f"Звіт_{safe_name}.xlsx"
    except:
        output_name = "Звіт_Ready.xlsx"

    output_path = os.path.join(output_dir, output_name)
    pd.DataFrame(processed_rows)[columns_list].to_excel(output_path, index=False)

    # ---------------------------------------------------------
    # ДИЗАЙН ЧЕРЕЗ OPENPYXL (МАГІЧНИЙ ФІКС ДЛЯ LIBREOFFICE)
    # ---------------------------------------------------------
    wb = load_workbook(output_path)
    ws = wb.active
    ws.row_dimensions[1].height = 80 # Шапка висока, щоб не ламало слова

    # 1. Повертаємо логіку ширин, яка гарантовано ловить колонки
    for col in ws.columns: 
        ws.column_dimensions[col[0].column_letter].width = 12

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    for idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        if 'піб' in header: ws.column_dimensions[col_letter].width = 35
        elif 'адрес' in header: ws.column_dimensions[col_letter].width = 45
        elif 'вулиц' in header or 'пункт' in header or 'район' in header: ws.column_dimensions[col_letter].width = 25
        elif 'рахун' in header or 'телефон' in header: ws.column_dimensions[col_letter].width = 16
        elif 'борг' in header or 'сплат' in header or 'заборг' in header or 'період' in header:
            ws.column_dimensions[col_letter].width = 16 # Зробив 16, щоб слово "Заборгованість" (14 літер) не переносилось по складах

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    city_idx = next((i + 1 for i, h in enumerate(headers) if 'населений пункт' in h), 1)

    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.font = Font(bold=True)
        cell.border = thin_border

    # 2. Форматуємо дані та ПРИМУСОВО розраховуємо висоту рядка
    for row in ws.iter_rows(min_row=2):
        city_val = row[city_idx - 1].value
        is_tot = "ВСЬОГО" in str(city_val).upper() if city_val else False
        is_empty_row = not bool(city_val)

        max_lines = 1 # Рахуємо кількість ліній для висоти цього конкретного рядка

        for cell in row:
            header = headers[cell.column - 1]

            if not is_empty_row:
                cell.border = thin_border

            # Вирівнювання
            if 'піб' in header or 'адрес' in header or 'вулиц' in header or 'пункт' in header:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
            else:
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

            if is_tot:
                cell.fill = yellow_fill
                cell.font = Font(bold=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

            # --- МАТЕМАТИКА ДЛЯ ВИСОТИ РЯДКА (Фікс для LibreOffice Calc) ---
            if cell.value and isinstance(cell.value, str):
                col_width = ws.column_dimensions[cell.column_letter].width
                # Скільки символів влазить в 1 рядок (мінус 1 на відступи)
                chars_per_line = col_width - 1
                lines_needed = math.ceil(len(cell.value) / max(chars_per_line, 1))
                
                if lines_needed > max_lines:
                    max_lines = lines_needed

        # Задаємо висоту рядка (1 рядок = ~15 пікселів)
        if max_lines > 1:
            # Обмежуємо максимум 90 пікселів, щоб не розтягнуло на пів екрану
            ws.row_dimensions[row[0].row].height = min(15 * max_lines, 90) 

    ws.freeze_panes = 'A2'
    wb.save(output_path)

    return output_name